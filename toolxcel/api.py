import openpyxl
from openpyxl.styles import Font, PatternFill
EXCEL_COMMON_TYPES = int|float|str

WORKBOOKPATH = ""

wb = openpyxl.load_workbook(WORKBOOKPATH)
ws = wb.active

def cell_to_coords(cell):
    col, row = '', ''
    for char in cell:
        if char.isdigit():
            row += char
        else:
            col += char
    col = sum((ord(char.lower()) - ord('a') + 1) * (26 ** i) for i, char in enumerate(reversed(col)))
    return (col, int(row))

# below are basic APIs

def set_text(dest: str, value: EXCEL_COMMON_TYPES):
    """
    Sets the text value in the specified destination cell.

    Args:
        dest (str): The destination cell address.
        value (EXCEL_COMMON_TYPES): The text value to be set.

    Returns:
        None
    """
    ws[dest] = value

def insert_row_before(dest: str):
    """
    Inserts a row before the specified destination cell.

    Args:
        dest (str): The destination cell reference.

    Returns:
        None
    """
    ws.insert_rows(cell_to_coords(dest)[0])

def insert_column_before(dest: str):
    """
    Inserts a column before the specified destination column.

    Args:
        dest (str): The destination column reference.

    Returns:
        None
    """
    ws.insert_cols(cell_to_coords(dest)[1])

# TODO: deletion

def set_title(text: str):
    """
    Sets the title of the worksheet.

    Args:
        text (str): The text to set as the title.
    """
    ws.title = text

def create_sheet(title: str):
    """
    Create a new sheet with the given title.

    Args:
        title (str): The title of the sheet.

    Returns:
        None
    """
    ws = wb.create_sheet(title)

def copy_region(dest: str, source: str):
    """
    Copy the contents of the source cell to the destination cell.

    Args:
        dest (str): The destination cell address.
        source (str): The source cell address.
    """
    ws[dest] = ws[source]

# above are basic APIs
    
# below are Excel functions

def f_sum(dest: str, regions: str):
    """
    Calculates the sum of the specified regions and stores the result in the specified destination cell.

    Args:
        dest (str): The destination cell where the sum will be stored.
        regions (str): The regions to be summed, separated by commas.

    Returns:
        None
    """
    ws[dest] = f"=SUM({regions})"

def f_average(dest: str, regions: str):
    """
    Calculate the average of the specified regions and store the result in the destination cell.

    Args:
        dest (str): The destination cell where the average will be stored.
        regions (str): The regions to calculate the average from.

    Returns:
        None
    """
    ws[dest] = f"=AVERAGE({regions})"

def f_max(dest: str, regions: str):
    """
    Calculates the maximum value from a range of cells and assigns it to the specified destination cell.

    Args:
        dest (str): The destination cell where the maximum value will be assigned.
        regions (str): The range of cells from which the maximum value will be calculated.

    Returns:
        None
    """
    ws[dest] = f"=MAX({regions})"

def f_min(dest: str, regions: str):
    """
    Calculates the minimum value from the specified regions and assigns it to the destination cell.

    Args:
        dest (str): The destination cell where the minimum value will be assigned.
        regions (str): The regions from which the minimum value will be calculated.

    Returns:
        None
    """
    ws[dest] = f"=MIN({regions})"

def f_count(dest: str, regions: str):
    """
    Calculates the count of cells in the specified regions and assigns the result to the destination cell.

    Args:
        dest (str): The destination cell where the count result will be assigned.
        regions (str): The regions to count cells from.

    Returns:
        None
    """
    ws[dest] = f"=COUNT({regions})"

def f_median(dest: str, regions: str):
    """
    Calculates the median of the specified regions and assigns the result to the destination cell.

    Args:
        dest (str): The destination cell where the result will be assigned.
        regions (str): The regions to calculate the median from.

    Returns:
        None
    """
    ws[dest] = f"=MEDIAN({regions})"

def f_stdev(dest: str, regions: str):
    """
    Calculates the standard deviation of the specified regions and stores the result in the destination cell.

    Parameters:
    dest (str): The destination cell where the result will be stored.
    regions (str): The regions to calculate the standard deviation from.

    Returns:
    None
    """
    ws[dest] = f"=STDEV({regions})"

def f_var(dest: str, regions: str):
    """
    Calculates the VAR (Variance) of the specified regions and assigns the result to the destination cell.

    Args:
        dest (str): The destination cell where the result will be assigned.
        regions (str): The regions to calculate the variance for.

    Returns:
        None
    """
    ws[dest] = f"=VAR({regions})"

def f_if(dest: str, condition: str, t_branch: EXCEL_COMMON_TYPES, f_branch: EXCEL_COMMON_TYPES):
    """
    A function that writes an IF formula to a specified destination cell in Excel.

    Parameters:
    dest (str): The destination cell where the formula will be written.
    condition (str): The condition for the IF formula.
    t_branch (EXCEL_COMMON_TYPES): The value or formula to be returned if the condition is true.
    f_branch (EXCEL_COMMON_TYPES): The value or formula to be returned if the condition is false.
    """
    ws[dest] = f"=IF({condition}, {t_branch}, {f_branch})"

def f_vlookup():
    pass

# above are Excel functions

# below are style functions

def set_font_style(dest: str, name: str, size: int|float, bold: bool, italic: bool, color: str):
    """
    Set the font style for the specified destination.

    Args:
        dest (str): The destination to set the font style for.
        name (str): The name of the font.
        size (int|float): The size of the font.
        bold (bool): Whether the font should be bold or not.
        italic (bool): Whether the font should be italic or not.
        color (str): The color of the font.

    Returns:
        None
    """
    ft = Font(name=name, size=size, bold=bold, italic=italic, color=color)
    # TODO

def set_background_style(dest: str, color: bool):
    """
    Sets the background style of the specified destination.

    Args:
        dest (str): The destination to set the background style for.
        color (bool): The color to set as the background style.

    Returns:
        None
    """
    fill = PatternFill(None, start_color=color, end_color=color)
    # TODO

# above are style functions
    
# below are chart functions
    
def create_bar_chart(dest: str, regions: str, containing_headers: bool):
    """
    Create a bar chart.

    Args:
        dest (str): The destination file path to save the bar chart.
        regions (str): The regions to include in the bar chart.
        containing_headers (bool): Whether the data contains headers.

    Returns:
        None
    """
    # TODO
    pass

# above are chart functions

# below are filter functions

def create_filter(region: str):
    """
    Create a filter based on the specified region.

    Args:
        region (str): The region to create the filter for.

    Returns:
        None
    """
    # TODO
    pass

def create_sort(region: str, ):
    """
    Create a sort in the specified region.

    Args:
        region (str): The region to apply the sort.

    Returns:
        None
    """
    # TODO
    # not clear in openpyxl doc
    pass

# above are filter functions
